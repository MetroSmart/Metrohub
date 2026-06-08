from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.export.base import ExportadorReporte

_AZUL_OSC  = "042C53"
_AZUL_MED  = "185FA5"
_VERDE     = "EAF3DE"
_WARN      = "FAEEDA"
_DANGER    = "FCEBEB"
_GRIS      = "F4F6F8"
_BORDE_CLR = "CCCCCC"

_ESTADO_FILL = {
    "operativo":     _VERDE,
    "aprobada":      _VERDE,
    "activo":        _VERDE,
    "mantenimiento": _WARN,
    "revision":      _WARN,
    "borrador":      _GRIS,
    "baja":          _DANGER,
    "reparacion":    _WARN,
}


def _border() -> Border:
    s = Side(style="thin", color=_BORDE_CLR)
    return Border(left=s, right=s, top=s, bottom=s)


def _header_cell(cell, texto: str, color: str = _AZUL_OSC) -> None:
    cell.value = texto
    cell.font = Font(bold=True, color="FFFFFF", size=9)
    cell.fill = PatternFill("solid", fgColor=color)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = _border()


def _data_cell(cell, valor, bg: str = "FFFFFF") -> None:
    cell.value = valor
    cell.font = Font(size=9)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    cell.border = _border()


def _portada(ws, fecha: str) -> None:
    ws.title = "Portada"
    ws.merge_cells("A1:D3")
    c = ws["A1"]
    c.value = "MetroHub — Informe Operativo ATU"
    c.font = Font(bold=True, color="FFFFFF", size=16)
    c.fill = PatternFill("solid", fgColor=_AZUL_OSC)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 30
    ws.row_dimensions[3].height = 30
    ws.merge_cells("A4:D4")
    ws["A4"].value = f"Metropolitano de Lima  ·  Fecha de generación: {fecha}"
    ws["A4"].font = Font(size=10, color=_AZUL_MED)
    ws["A4"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[4].height = 20
    for col in range(1, 5):
        ws.column_dimensions[get_column_letter(col)].width = 22


def _sheet_tabla(
    wb: Workbook,
    titulo: str,
    encabezados: list[str],
    filas: list[list],
    anchos: list[int],
    col_estado_idx: int | None = None,
    header_color: str = _AZUL_OSC,
) -> None:
    ws = wb.create_sheet(titulo)
    for col_i, enc in enumerate(encabezados, start=1):
        _header_cell(ws.cell(row=1, column=col_i), enc, header_color)
    ws.row_dimensions[1].height = 18

    for row_i, fila in enumerate(filas, start=2):
        bg_row = _GRIS if row_i % 2 == 0 else "FFFFFF"
        for col_i, val in enumerate(fila, start=1):
            bg = bg_row
            if col_estado_idx is not None and col_i == col_estado_idx + 1:
                bg = _ESTADO_FILL.get(str(val), bg_row)
            _data_cell(ws.cell(row=row_i, column=col_i), val, bg=bg)
        ws.row_dimensions[row_i].height = 16

    for col_i, ancho in enumerate(anchos, start=1):
        ws.column_dimensions[get_column_letter(col_i)].width = ancho

    if not filas:
        ws.merge_cells(f"A2:{get_column_letter(len(encabezados))}2")
        ws["A2"].value = "Sin registros."
        ws["A2"].font = Font(italic=True, color="888888", size=9)


class XlsxExporter(ExportadorReporte):
    def exportar(self, datos: dict[str, Any]) -> tuple[bytes, str, str]:
        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

        _portada(wb.create_sheet("Portada"), datos.get("fecha", ""))

        # Resumen Ejecutivo
        _KPI_LABELS = {
            "rutas_activas":         "Rutas activas",
            "choferes_activos":      "Choferes activos",
            "buses_operativos":      "Buses operativos",
            "asignaciones_hoy":      "Asignaciones confirmadas hoy",
            "conflictos_abiertos":   "Conflictos abiertos",
            "certif_por_vencer_30d": "Certificaciones por vencer (30 días)",
        }
        kpis = datos.get("kpis", {})
        _sheet_tabla(wb, "Resumen",
            ["Indicador", "Valor"],
            [[_KPI_LABELS.get(k, k.replace("_", " ").capitalize()), v]
             for k, v in kpis.items()],
            [38, 14],
        )

        # Rutas Activas
        _sheet_tabla(wb, "Rutas Activas",
            ["Código", "Nombre", "Tipo", "Hora Inicio", "Hora Fin", "Frec. (min)"],
            [[r["codigo"], r["nombre"], r["tipo"],
              r["hora_inicio"], r["hora_fin"], r["frecuencia_min"]]
             for r in datos.get("rutas", [])],
            [10, 30, 14, 12, 12, 13],
        )

        # Flota
        _sheet_tabla(wb, "Flota",
            ["Placa", "Tipo", "Año", "Capacidad", "Estado"],
            [[b["placa"], b["tipo"], b["anio"], b["capacidad"], b["estado"]]
             for b in datos.get("buses", [])],
            [12, 16, 8, 12, 16],
            col_estado_idx=4,
        )

        # Choferes
        _sheet_tabla(wb, "Choferes",
            ["Nombre", "Tipo Licencia", "N° Licencia", "Vence Licencia", "Vence Certif."],
            [[c["nombre"], c["licencia"], c["numero"], c["vence_lic"], c["vence_certif"]]
             for c in datos.get("choferes", [])],
            [28, 14, 16, 16, 16],
        )

        # Alertas de Documentos
        _sheet_tabla(wb, "Alertas Documentos",
            ["Chofer", "Vence Licencia", "Días Rest.", "Vence Certif.", "Días Rest."],
            [[a["nombre"], a["vence_lic"], a["dias_lic"], a["vence_certif"], a["dias_certif"]]
             for a in datos.get("alertas_doc", [])],
            [28, 16, 14, 16, 14],
            header_color="854F0B",
        )
        ws_alerta = wb["Alertas Documentos"]
        for row in ws_alerta.iter_rows(min_row=2, max_col=5):
            if row[0].value is None:
                break
            dias_lic = row[2].value
            dias_cer = row[4].value
            if isinstance(dias_lic, int) and isinstance(dias_cer, int):
                bg = _DANGER if (dias_lic <= 7 or dias_cer <= 7) else _WARN
                for cell in row:
                    cell.fill = PatternFill("solid", fgColor=bg)

        # Programaciones
        _sheet_tabla(wb, "Programaciones",
            ["Nombre", "Estado", "Fecha Inicio", "Fecha Fin"],
            [[p["nombre"], p["estado"], p["inicio"], p["fin"]]
             for p in datos.get("programaciones", [])],
            [28, 14, 14, 14],
            col_estado_idx=1,
        )

        # Conflictos
        _sheet_tabla(wb, "Conflictos",
            ["Tipo", "Descripción"],
            [[c["tipo"], c["descripcion"]] for c in datos.get("conflictos", [])],
            [22, 50],
            header_color="A32D2D",
        )

        buf = BytesIO()
        wb.save(buf)
        return (
            buf.getvalue(),
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "xlsx",
        )
